#!/usr/bin/env python3
"""Pipeline complet : un PDF déposé dans Data/ → classeur + page à jour.

    python3 pipeline/pipeline.py --dossier .
    python3 pipeline/pipeline.py --dossier . --simulation   (ne rien écrire)

Détecte les sondages non encore intégrés, les extrait, les saisit, recalcule,
passe les contrôles et met la page à jour. S'arrête bruyamment au moindre doute :
mieux vaut ne rien faire que produire une projection fausse.

CONVENTIONS
  · Date       : fin de la période de terrain, lue DANS le rapport (jamais dans
                 le nom de fichier, qui porte la date de publication).
  · n          : la colonne que la firme met en tête de son tableau de régions —
                 pour Léger la sous-colonne « décidés », pour Synopsis et Pallas
                 le total. La même valeur va dans le classeur ET dans le CSV.
  · Régions    : les n régionaux sont LUS dans le rapport, jamais estimés.
"""
import argparse, glob, io, os, re, subprocess, sys, shutil, tempfile
from datetime import date, timedelta

PARTIS = ['CAQ','PLQ','QS','PQ','PCQ','Autre']
CSV_PARTIS = ['CAQ','PLQ','QS','PQ','PCQ']   # ordre du fichier data/sondages_national.csv
MOIS = ['janvier','février','mars','avril','mai','juin','juillet','août',
        'septembre','octobre','novembre','décembre']
MOIS_RE = ('janvier|f[ée]vrier|mars|avril|mai|juin|juillet|ao[uû]t|'
           'septembre|octobre|novembre|d[ée]cembre')

def normaliser(s):
    """Ramène « Léger », « leger », « Leger » à une même clé. Le classeur contient
    historiquement plusieurs graphies : on compare sur une forme canonique."""
    import unicodedata
    s = unicodedata.normalize('NFD', str(s or '')).encode('ascii', 'ignore').decode().lower()
    return re.sub(r'[^a-z]', '', s)

def num_mois(nom):
    n = normaliser(nom)
    for i, m in enumerate(MOIS):
        if normaliser(m).startswith(n[:4]):
            return i + 1
    return None

class Arret(Exception):
    """Condition d'arrêt : on remonte le cas plutôt que de deviner."""

# ─────────────────────────── extraction des PDF ───────────────────────────

def nombres(ligne):
    return [float(x.replace(',', '.')) for x in re.findall(r'(\d+(?:[.,]\d+)?)%', ligne)]

def cases(ligne):
    """Découpe une ligne de tableau en cellules. Les milliers sont séparés par une
    espace simple (« 1 008 ») : on ne coupe donc que sur deux espaces ou plus."""
    return [c for c in re.split(r'\s{2,}', ligne.strip()) if re.search(r'\d', c)]

MOTIFS = {
    'PQ':    r'(?:Le )?Parti [Qq]uéb[ée]cois|^PQ,',
    'PLQ':   r'(?:Le )?Parti lib[ée]ral|^PLQ,',
    'CAQ':   r'(?:La )?Coalition [Aa]venir|^CAQ,',
    'PCQ':   r'(?:Le )?Parti conservateur|^(?:PCQ|Conservateur),',
    'QS':    r'Qu[ée]bec solidaire|^QS,',
    'Autre': r'Un autre parti|^Autre Parti',
}
# Léger publie une colonne « décidés » qui somme déjà à 100. Synopsis et Pallas
# publient les indécis sur une ligne distincte : leurs partis somment à ~84, et
# c'est normal. On capte donc les indécis pour valider partis + indécis ≈ 100.
MOTIF_INDECIS = r'^Ind[ée]cis'

COLONNES = {
    # firme : marqueur de la ligne des effectifs, index de la colonne provinciale,
    #         index des colonnes régionales, nombre minimal de colonnes du tableau.
    # Le marqueur seul ne suffit pas : plusieurs tableaux d'un même rapport le
    # portent. On retient la première ligne qui a AUSSI le bon nombre de colonnes.
    'leger':    dict(n=r'n absolu=', prov=1,
                     reg={'Montréal RMR': 9, 'Québec RMR': 10, 'Reste du Québec': 11}, mini=13),
    'synopsis': dict(n=r'^\s*n=', prov=0,
                     reg={'Île de Montréal': 8, 'Banlieue de Montréal': 9,
                          'Québec RMR': 10, 'Reste du Québec': 11}, mini=13),
    'pallas':   dict(n=r'^Fr[ée]quence(?! pond)', prov=0,
                     reg={'Montréal RMR': 7, 'Île de Montréal': 8, 'Banlieue de Montréal': 9,
                          'Québec RMR': 10, 'Reste du Québec': 11}, mini=14),
}

def date_terrain(txt, indice, fichier):
    """Date de référence du sondage : le MILIEU de la période de collecte, lue dans
    le rapport. Renvoie (date ISO, description de la période).

    Deux raisons de ne pas se fier au nom de fichier. D'abord il porte la date de
    publication : Léger publie le 1er septembre un sondage mené du 28 au 31 août ;
    s'y fier crée un doublon décalé de plusieurs jours, invisible à l'œil dans le
    classeur. Ensuite un sondage ne mesure pas un instant mais un intervalle, et le
    milieu de l'intervalle en est le meilleur repère — c'est la convention usuelle
    des agrégateurs. Une période de deux à cinq jours ne déplace le repère que d'un
    ou deux jours, négligeable devant la fenêtre de 25 jours du lissage, mais la
    CONSTANCE, elle, compte.

    Quatre formes rencontrées :
        « 28 au 31 août 2026 »            (Léger)
        « Période : 24 au 26 août 2026 »  (Synopsis)
        « Du 1er au 2 août 2026 »         (Pallas, période)
        « Les 10 et 12 juin 2026 »        (Pallas, deux journées)
        « Le 29 août 2026, Pallas… »      (Pallas, journée unique)
    """
    tete = '\n'.join(txt.split('\n')[:400])
    bornes = None
    # a) période « X [mois] au Y mois AAAA »
    m = re.search(r'(\d{1,2})\s*(?:er)?\s*(' + MOIS_RE + r')?\s*au\s+'
                  r'(\d{1,2})\s*(?:er)?\s+(' + MOIS_RE + r')\s+(\d{4})', tete, re.I)
    if m:
        an, mo2 = int(m.group(5)), num_mois(m.group(4))
        mo1 = num_mois(m.group(2)) if m.group(2) else mo2
        bornes = ((int(m.group(1)), mo1), (int(m.group(3)), mo2), an)
    if not bornes:
        # b) deux journées « Les X et Y mois AAAA »
        m = re.search(r'\b[Ll]es\s+(\d{1,2})\s+et\s+(\d{1,2})\s*(?:er)?\s+('
                      + MOIS_RE + r')\s+(\d{4})', tete, re.I)
        if m:
            mo, an = num_mois(m.group(3)), int(m.group(4))
            bornes = ((int(m.group(1)), mo), (int(m.group(2)), mo), an)
    if not bornes:
        # c) journée unique « Le X mois AAAA »
        m = re.search(r'\b[Ll]e\s+(\d{1,2})\s*(?:er)?\s+(' + MOIS_RE + r')\s+(\d{4})', tete, re.I)
        if m:
            mo, an = num_mois(m.group(2)), int(m.group(3))
            bornes = ((int(m.group(1)), mo), (int(m.group(1)), mo), an)
    if not bornes:
        raise Arret(f"{fichier} : période de terrain introuvable dans le rapport. "
                    f"La date du nom de fichier est la date de PUBLICATION et ne doit "
                    f"pas servir : elle décale le sondage de plusieurs jours et crée un "
                    f"doublon. Saisir ce sondage à la main.")
    (j1, m1), (j2, m2), an = bornes
    if not (m1 and m2):
        raise Arret(f"{fichier} : mois illisible dans la période de terrain.")
    try:
        d1 = date(an, m1, j1)
        d2 = date(an, m2, j2)
        if d2 < d1:                      # période à cheval sur le 31 décembre
            d2 = date(an + 1, m2, j2)
    except ValueError:
        raise Arret(f"{fichier} : date de terrain invalide ({j1}/{m1} – {j2}/{m2}/{an}).")
    if not (0 <= (d2 - d1).days <= 21):
        raise Arret(f"{fichier} : période de terrain de {(d2-d1).days} jours "
                    f"({d1} → {d2}) — invraisemblable, la phrase captée n'est pas la bonne.")
    # milieu, arrondi au jour supérieur : une période de deux jours prend le second
    d = d1 + timedelta(days=((d2 - d1).days + 1) // 2)
    ecart = abs((d - indice).days)
    if ecart > 30:
        raise Arret(f"{fichier} : la période lue ({d1} → {d2}) est à {ecart} jours "
                    f"de la date du nom de fichier ({indice.isoformat()}). C'est trop : la "
                    f"phrase captée n'est probablement pas la bonne. Vérification manuelle.")
    periode = d1.isoformat() if d1 == d2 else f"{d1.isoformat()} → {d2.isoformat()}"
    return d.isoformat(), periode

def ligne_effectifs(txt, spec, fichier):
    """Effectifs du tableau régional. Plusieurs tableaux d'un rapport portent le même
    marqueur (« Fréquence non pondérée » apparaît sous chaque croisement) : on retient
    la première ligne qui a le nombre de colonnes du tableau régional."""
    for ligne in txt.split('\n'):
        cible = ligne.strip() if spec['n'].startswith('^') else ligne
        if not re.search(spec['n'], cible, re.I):
            continue
        reste = re.sub(spec['n'].lstrip('^').replace(r'\s*', ''), '', ligne, count=1, flags=re.I)
        c = cases(reste)
        if len(c) >= spec['mini']:
            return [int(re.sub(r'\D', '', x) or 0) for x in c]
    raise Arret(f"{fichier} : ligne des effectifs (n) introuvable, ou trop courte pour le "
                f"tableau régional ({spec['mini']} colonnes attendues). Le rapport a "
                f"probablement changé de mise en page.")

def extraire(pdf, firme):
    fichier = os.path.basename(pdf)
    if firme not in COLONNES:
        raise Arret(f"firme inconnue : « {firme} ». Les formats reconnus sont "
                    f"{', '.join(COLONNES)}. Un nouveau format doit être ajouté à la main "
                    f"après vérification visuelle du rapport.")
    txt = subprocess.run(['pdftotext', '-layout', pdf, '-'],
                         capture_output=True, text=True, timeout=120).stdout
    if re.search(r'grandes tendances r[ée]gionales|a cumul[ée]\s+\d+\s+sondages', txt, re.I):
        raise Arret(f"{fichier} est un rapport CUMULATIF (plusieurs sondages agrégés). "
                    f"Ses tableaux régionaux sont des graphiques, pas du texte : l'identité des partis "
                    f"s'y lit aux couleurs. Il sert aux décalages régionaux, pas au niveau national, "
                    f"et doit être saisi à la main après vérification visuelle.")
    if not txt.strip():
        raise Arret(f"{fichier} : aucun texte extrait. PDF scanné, ou tableaux "
                    f"sous forme d'images — dans ce cas l'identité des partis se lit aux couleurs "
                    f"et l'erreur serait silencieuse. Traitement manuel obligatoire.")
    spec = COLONNES[firme]
    brut = {}
    for parti, motif in MOTIFS.items():
        for ligne in txt.split('\n'):
            if re.search(motif, ligne.strip(), re.I):
                v = nombres(ligne)
                if len(v) >= spec['mini']:
                    brut[parti] = v
                    break
    indecis = None
    for ligne in txt.split('\n'):
        if re.search(MOTIF_INDECIS, ligne.strip(), re.I):
            v = nombres(ligne)
            if len(v) >= spec['mini']:
                indecis = v
                break
    manquants = [p for p in PARTIS if p not in brut]
    if manquants:
        raise Arret(f"{fichier} : partis introuvables dans le tableau — {manquants}. "
                    f"Le rapport a probablement changé de mise en page.")

    eff = ligne_effectifs(txt, spec, fichier)
    out, effectifs = {}, {}
    for nom, idx in [('PROVINCE', spec['prov'])] + list(spec['reg'].items()):
        v = [brut[p][idx] for p in PARTIS]
        total = sum(v) + (indecis[idx] if indecis else 0)
        if not (96 <= total <= 104):
            quoi = "partis + indécis" if indecis else "parts"
            raise Arret(f"{fichier} / {nom} : {quoi} somment à {total:.0f}, "
                        f"pas à 100. Colonne mal lue — la mise en page du rapport a peut-être changé.")
        s = sum(v)
        out[nom] = [x * 100 / s for x in v]   # répartition proportionnelle des indécis
        if idx >= len(eff):
            raise Arret(f"{fichier} / {nom} : pas d'effectif en colonne {idx}.")
        effectifs[nom] = eff[idx]

    # Un sous-échantillon régional ne peut pas dépasser l'échantillon provincial,
    # et la somme des régions disjointes doit s'en approcher. Si ce n'est pas le
    # cas, c'est qu'on lit la ligne d'effectifs d'un autre tableau.
    np_ = effectifs['PROVINCE']
    if np_ < 300:
        raise Arret(f"{fichier} : effectif provincial lu à {np_}, invraisemblable.")
    for nom, e in effectifs.items():
        if nom != 'PROVINCE' and not (30 <= e <= np_):
            raise Arret(f"{fichier} / {nom} : effectif régional lu à {e}, hors de [30, {np_}]. "
                        f"La ligne des effectifs ne correspond pas au tableau régional.")
    # « Montréal RMR » recouvre l'Île et sa banlieue : il n'est redondant que si les
    # deux moitiés sont publiées à part. Chez Léger, qui ne les sépare pas, c'est au
    # contraire l'une des trois régions disjointes.
    detail_mtl = {'Île de Montréal', 'Banlieue de Montréal'} <= set(effectifs)
    disjointes = [n for n in effectifs
                  if n != 'PROVINCE' and not (detail_mtl and n == 'Montréal RMR')]
    somme = sum(effectifs[n] for n in disjointes)
    if disjointes and abs(somme - np_) > max(0.15 * np_, 60):
        raise Arret(f"{fichier} : les régions disjointes ({', '.join(disjointes)}) somment à "
                    f"{somme} pour un échantillon de {np_}. Les colonnes lues ne sont pas "
                    f"les bonnes.")

    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', fichier)
    indice = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    d, periode = date_terrain(txt, indice, fichier)
    return out, effectifs, d, periode

# ─────────────────────────── saisie dans le classeur ───────────────────────────

def reancrer(formule, source, cible):
    """Recopie une formule d'une ligne à une autre en ne touchant QUE les numéros de
    ligne des références de cellules relatives.

    Le piège : une substitution naïve sur « le chiffre 4 » transforme
    LEFT(A4,4) en LEFT(A4,63) — l'argument de longueur est massacré et la formule
    renvoie n'importe quoi, sans erreur visible. On ne remplace donc que ce qui
    suit immédiatement une lettre de colonne, et jamais une ligne absolue ($4)."""
    def repl(m):
        dollar_col, col, dollar_lig, lig = m.groups()
        if dollar_lig == '$' or int(lig) != source:
            return m.group(0)
        return f"{dollar_col}{col}{dollar_lig}{cible}"
    return re.sub(r'(\$?)([A-Z]{1,3})(\$?)(\d+)', repl, formule)

# Plages qui balaient la feuille Sondages, sous leurs deux formes :
#   · depuis une autre feuille   →  Sondages!$M$4:$M$62
#   · depuis la feuille Sondages →  E$4:E$62,  $K$4:$K$62
# Le lookbehind (?<![!$A-Z0-9]) empêche de confondre la seconde forme avec une plage
# d'une AUTRE feuille citée depuis Sondages (Modèle!$D$2:$D$62).
PLAGE_EXT = re.compile(r'(Sondages!\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?)(\d+)(?![\d(])')
PLAGE_INT = re.compile(r'(?<![!$A-Z0-9])(\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?)(\d+)(?![\d(])')

def etendre_plages(wb, ancienne, nouvelle):
    """Recale sur la nouvelle dernière ligne les plages qui balaient Sondages.

    Les formules sont écrites en dur (Sondages!$M$4:$M$62). Une ligne ajoutée
    au-delà tombe hors de la fenêtre de régression : le sondage est dans le
    classeur, visible, mais le modèle ne le voit pas. Les sièges ne bougent pas
    d'un siège et tout a l'air normal — c'est le défaut le plus dangereux du lot.
    Symétriquement, une plage plus longue que la feuille fait entrer des cellules
    vides dans un SUMPRODUCT, qui renvoie alors #VALEUR!. On ne remplace donc que
    les plages qui se terminaient EXACTEMENT à l'ancienne dernière ligne : une
    plage qui vise autre chose n'est pas touchée."""
    def sub(motif, texte):
        return motif.sub(lambda m: m.group(1) + (str(nouvelle) if m.group(2) == str(ancienne)
                                                 else m.group(2)), texte)
    touchees = 0
    for ws in wb.worksheets:
        for rangee in ws.iter_rows():
            for c in rangee:
                if not (isinstance(c.value, str) and c.value.startswith('=')): continue
                neuf = sub(PLAGE_EXT, c.value)
                if ws.title == 'Sondages':
                    neuf = sub(PLAGE_INT, neuf)
                if neuf != c.value:
                    c.value = neuf
                    touchees += 1
    return touchees

def plage_agregat(chemin):
    """Dernière ligne balayée par les formules, et dernière ligne réelle de la
    feuille Sondages. Les deux doivent coïncider."""
    import openpyxl
    wb = openpyxl.load_workbook(chemin)
    fins = set()
    for ws in wb.worksheets:
        for rangee in ws.iter_rows():
            for c in rangee:
                if not (isinstance(c.value, str) and c.value.startswith('=')): continue
                fins.update(int(m.group(2)) for m in PLAGE_EXT.finditer(c.value))
                if ws.title == 'Sondages':
                    fins.update(int(m.group(2)) for m in PLAGE_INT.finditer(c.value))
    fins = {f for f in fins if f > 10}      # on ignore les petites plages de tête
    return (min(fins) if fins else None), wb['Sondages'].max_row

def cle_de(ws, r):
    return f"{ws.cell(r,1).value}|{normaliser(ws.cell(r,2).value)}"

def doublon(ws, date_s, firme, parts_province):
    """Un même sondage peut arriver deux fois sous deux dates : le rapport de terrain
    et le rapport régional publié quelques jours plus tard. La clé date+firme ne suffit
    donc pas. On refuse aussi un sondage de la même firme, à moins de six jours, dont
    les parts provinciales sont identiques."""
    d0 = date(*map(int, date_s.split('-')))
    f0 = normaliser(firme)
    for r in range(4, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v or normaliser(ws.cell(r, 2).value) != f0: continue
        if str(ws.cell(r, 3).value) != 'PROVINCE': continue
        try:
            d1 = date(*map(int, str(v)[:10].split('-')))
        except ValueError:
            continue
        if abs((d1 - d0).days) > 6: continue
        anciennes = [ws.cell(r, 5 + i).value for i in range(5)]
        if any(x is None for x in anciennes): continue
        if all(abs(float(a) - b) <= 0.6 for a, b in zip(anciennes, parts_province[:5])):
            return str(v)[:10]
    return None

def saisir(classeur, sortie, nouveaux):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.load_workbook(classeur)
    ws = wb['Sondages']
    bleu = Font(color='0000CC')
    depart = ws.max_row             # dernière ligne AVANT l'ajout : c'est elle que
    ligne = depart + 1              # les plages en dur visent aujourd'hui
    modele = 4                      # ligne dont on recopie les formules
    ajouts, retenus = 0, []
    for date_s, firme, eff, parts in nouveaux:
        if any(cle_de(ws, r) == f"{date_s}|{normaliser(firme)}" for r in range(4, ws.max_row + 1)):
            print(f"  · {date_s} {firme} : déjà présent, ignoré")
            continue
        jumeau = doublon(ws, date_s, firme, parts['PROVINCE'])
        if jumeau:
            print(f"  · {date_s} {firme} : mêmes chiffres que la ligne du {jumeau} "
                  f"(même firme, moins de six jours) — doublon, ignoré")
            continue
        for region, v in parts.items():
            ws.cell(ligne, 1, date_s); ws.cell(ligne, 2, firme); ws.cell(ligne, 3, region)
            ws.cell(ligne, 4, eff[region])
            for i in range(6):
                c = ws.cell(ligne, 5 + i, round(v[i], 1)); c.font = bleu
            for col in range(11, 21):
                f = ws.cell(modele, col).value
                if isinstance(f, str) and f.startswith('='):
                    ws.cell(ligne, col, reancrer(f, modele, ligne))
            ligne += 1; ajouts += 1
        retenus.append((date_s, firme, eff, parts))
    if not ajouts:
        return None, [], 0
    touchees = etendre_plages(wb, depart, ws.max_row)
    wb.save(sortie)
    return ajouts, retenus, touchees

def ajouter_au_csv(chemin, retenus):
    """Le classeur et la page ne lisent pas la même source : le classeur a sa feuille
    Sondages, la page lit ce CSV. Ne remplir que l'un des deux, c'est faire diverger
    les sièges et la courbe sans que rien ne le signale."""
    texte = io.open(chemin, encoding='utf-8').read()
    deja = {l.split('|')[0] + '|' + normaliser(l.split('|')[1])
            for l in texte.strip().split('\n')[1:] if '|' in l}
    lignes = []
    for date_s, firme, eff, parts in retenus:
        if f"{date_s}|{normaliser(firme)}" in deja: continue
        v = parts['PROVINCE']
        lignes.append('|'.join([date_s, firme, str(eff['PROVINCE'])]
                               + [f"{v[PARTIS.index(p)]:.0f}" for p in CSV_PARTIS]))
    if not lignes:
        return 0
    if not texte.endswith('\n'):
        texte += '\n'               # le fichier livré n'a pas de retour final
    io.open(chemin, 'w', encoding='utf-8').write(texte + '\n'.join(lignes) + '\n')
    return len(lignes)

# ─────────────────────────── contrôles ───────────────────────────

def recalculer(chemin):
    if not shutil.which('soffice'):
        raise Arret("LibreOffice absent : impossible de recalculer le classeur.")
    d = tempfile.mkdtemp()
    subprocess.run(['soffice', '--headless', '--convert-to', 'xlsx', '--outdir', d, chemin],
                   capture_output=True, timeout=300)
    out = os.path.join(d, os.path.basename(chemin))
    if not os.path.exists(out):
        raise Arret("le recalcul par LibreOffice a échoué.")
    return out

def niveau_national(calcule):
    """Niveau national calculé par le classeur. Renvoie None si une case n'est pas
    un nombre : LibreOffice écrit « #VALUE! » quand une plage déborde de la feuille,
    et il vaut mieux le dire que comparer une chaîne à un nombre."""
    import openpyxl
    ag = openpyxl.load_workbook(calcule, data_only=True)['Agrégat']
    v = [ag.cell(17, c).value for c in range(2, 7)]
    return v if all(isinstance(x, (int, float)) for x in v) else None

def controles(calcule, source, precedent=None, avant=None):
    import openpyxl, numpy as np, collections
    wb = openpyxl.load_workbook(calcule, data_only=True)
    md, sg = wb['Modèle'], wb['Sièges']
    V = np.array([[md.cell(r, 4 + k).value or 0 for k in range(6)] for r in range(2, 129)], float)
    reg = dict(collections.Counter(md.cell(r, 2).value for r in range(2, 129)))
    ech = []
    if len(V) != 127: ech.append(f"127 circonscriptions attendues, {len(V)} trouvées")
    if reg != {'Île de Montréal': 28, 'Banlieue de Montréal': 40,
               'Québec RMR': 13, 'Reste du Québec': 46}: ech.append(f"répartition régionale : {reg}")
    sieges = {}
    for r in range(6, 20):
        nom = sg.cell(r, 1).value
        if not nom or not str(nom).startswith('Scén'): continue
        v = [sg.cell(r, c).value for c in range(2, 8)]
        if any(x is None for x in v): continue
        if sum(v) != 127: ech.append(f"{nom} : {sum(v)} sièges au lieu de 127")
        sieges[re.sub(r'^Scén\.\s*', '', str(nom))] = dict(zip(PARTIS, [int(x) for x in v]))
    if abs(V.sum() - 4112821) > 200: ech.append(f"base de référence : {V.sum():,.0f} au lieu de 4 112 821")
    prov = V.sum(0) / V.sum() * 100
    if abs(prov[3] - 14.6) > 0.2 or prov[5] < 1.3:
        ech.append(f"référence provinciale suspecte : PQ {prov[3]:.2f}, Autre {prov[5]:.2f} "
                   f"(fuite « Autre → PQ » ?)")
    maxi = max(max(v.values()) for v in sieges.values()) if sieges else 0
    if maxi >= 64:
        ech.append(f"UN PARTI ATTEINT LA MAJORITÉ ({maxi}). La page affirme partout le contraire : "
                   f"elle doit être réécrite, pas rafraîchie.")
    if precedent:
        for p in PARTIS[:5]:
            d = sieges.get('A – Swing proportionnel', {}).get(p, 0) - precedent.get(p, 0)
            if abs(d) > 10:
                ech.append(f"{p} bouge de {d:+d} sièges — au-delà du seuil de 10, on s'arrête.")

    # 9e contrôle — le sondage ajouté doit être VU par le modèle.
    fin_plage, fin_feuille = plage_agregat(source)
    if fin_plage is None:
        ech.append("aucune plage « Sondages! » trouvée dans les formules : classeur inattendu.")
    elif fin_plage < fin_feuille:
        ech.append(f"les formules de l'Agrégat s'arrêtent à la ligne {fin_plage} alors que la "
                   f"feuille Sondages en compte {fin_feuille} : les lignes ajoutées sont "
                   f"INVISIBLES au modèle.")
    apres = niveau_national(calcule)
    if apres is None:
        ech.append("le niveau national de l'Agrégat ne renvoie pas un nombre (#VALEUR! ?) : "
                   "une plage de formule déborde de la feuille Sondages.")
    elif avant is not None:
        if all(abs(a - b) < 0.01 for a, b in zip(avant, apres)):
            ech.append("le niveau national n'a bougé d'aucun centième après l'ajout du sondage. "
                       "Un sondage sans le moindre effet signale une saisie hors plage, pas une "
                       "stabilité de l'opinion.")
    return sieges, ech

# ─────────────────────────── orchestration ───────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dossier', default='.')
    ap.add_argument('--classeur', help='par défaut : le ModeleVivantQC127 le plus récent')
    ap.add_argument('--csv', default='data/sondages_national.csv')
    ap.add_argument('--simulation', action='store_true', help='tout vérifier sans rien écrire')
    a = ap.parse_args()
    os.chdir(a.dossier)

    classeur = a.classeur or sorted(glob.glob('ModeleVivantQC127_v*.xlsx'))[-1]
    print(f"Classeur      : {classeur}")

    import openpyxl
    ws = openpyxl.load_workbook(classeur)['Sondages']
    connus = {cle_de(ws, r) for r in range(4, ws.max_row + 1) if ws.cell(r, 1).value}

    nouveaux, ignores = [], []
    for pdf in sorted(glob.glob('Data/*.pdf')):
        m = re.match(r'(\d{4}-\d{2}-\d{2})-([a-z]+)\.pdf$', os.path.basename(pdf))
        if not m:
            print(f"  ! {os.path.basename(pdf)} : nom non conforme (AAAA-MM-JJ-firme.pdf), ignoré")
            continue
        _, firme = m.groups()
        etiquette = {'leger': 'Léger', 'synopsis': 'Synopsis', 'pallas': 'Pallas'}.get(firme, firme)
        try:
            parts, eff, d, periode = extraire(pdf, firme)
        except Arret as e:
            if 'CUMULATIF' in str(e):
                print(f"  ~ {os.path.basename(pdf)} : rapport cumulatif, ignoré (saisie manuelle)")
                ignores.append(os.path.basename(pdf))
                continue
            raise
        if f"{d}|{normaliser(etiquette)}" in connus:
            continue
        jumeau = doublon(ws, d, etiquette, parts['PROVINCE'])
        if jumeau:
            print(f"  · {os.path.basename(pdf)} : terrain au {d}, mêmes chiffres que la ligne "
                  f"du {jumeau} — même sondage, ignoré")
            continue
        print(f"  → nouveau : {d} {etiquette} (n={eff['PROVINCE']}) — terrain {periode}, "
              f"fichier daté du {m.group(1)}")
        nouveaux.append((d, etiquette, eff, parts))

    if not nouveaux:
        print("\nAucun sondage nouveau. Rien à faire.")
        return 0

    print(f"\n{len(nouveaux)} sondage(s) à intégrer.")
    for d, f, eff, parts in nouveaux:
        print(f"  {d} {f}")
        for r, v in parts.items():
            print(f"    {r:24s} n={eff[r]:5d}  " + "  ".join(f"{p} {x:4.1f}"
                                                            for p, x in zip(PARTIS[:5], v)))

    if a.simulation:
        print("\n[simulation] rien n'a été écrit.")
        return 0

    avant = niveau_national(recalculer(classeur))

    n_ver = int(re.search(r'_v(\d+)', classeur).group(1))
    sortie = re.sub(r'_v\d+[^.]*', f'_v{n_ver}_{nouveaux[-1][0].replace("-","")}', classeur)
    ajouts, retenus, touchees = saisir(classeur, sortie, nouveaux)
    if not ajouts:
        print("Rien de neuf après vérification.")
        return 0
    print(f"\nClasseur écrit : {sortie}  ({ajouts} lignes, {touchees} formules ré-étendues)")

    calcule = recalculer(sortie)
    sieges, echecs = controles(calcule, sortie, avant=avant)
    print("\nSièges :")
    for k, v in sieges.items():
        print(f"  {k:42s} " + "  ".join(f"{p} {v[p]:3d}" for p in ['PQ','PLQ','CAQ','PCQ','QS']))

    if echecs:
        print("\n╔═══ ARRÊT ═══")
        for e in echecs: print("║ ✘ " + e)
        print(f"╚═ Le classeur {sortie} est écrit mais SUSPECT ; ni le CSV ni la page n'ont bougé.")
        return 2

    print("\nContrôles : tous passés ✔")
    ajoutees = ajouter_au_csv(a.csv, retenus)
    print(f"{a.csv} : {ajoutees} ligne(s) ajoutée(s)")

    r = subprocess.run([sys.executable, 'pipeline/maj_artefact.py', '--classeur', sortie, '--recalculer'],
                       capture_output=True, text=True)
    print(r.stdout or r.stderr)
    print("\nIl reste UNE chose à faire à la main : relire le texte de la page,")
    print("et republier l'artefact depuis une session Claude.")
    return 0

if __name__ == '__main__':
    try:
        sys.exit(main())
    except Arret as e:
        print(f"\n╔═══ ARRÊT ═══\n║ {e}\n╚═ Rien n'a été modifié.")
        sys.exit(2)
