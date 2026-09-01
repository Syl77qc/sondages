#!/usr/bin/env python3
"""Met à jour la page « Tendances Québec 2026 » à partir des données courantes.

    python3 maj_artefact.py --sondages data/sondages_national.csv \
                            --classeur ModeleVivantQC127_v8_langue.xlsx \
                            --page livrables/tendances.html

La source est le fragment livrables/tendances.html (sans squelette HTML) ; la copie
autonome tendances-quebec-2026.html, à la racine, est régénérée à partir de lui.

Recalcule le lissage, remplace les blocs de données de la page, met à jour les
dates, et signale les phrases du texte qui pourraient être devenues fausses.
Ne réécrit jamais le texte : c'est à un humain de le faire, en connaissance de cause.
"""
import argparse, io, json, re, sys
import numpy as np
from datetime import date, timedelta

P = ['CAQ','PLQ','QS','PQ','PCQ']
MOIS = ['janvier','février','mars','avril','mai','juin','juillet','août',
        'septembre','octobre','novembre','décembre']
LETTRES = {1:'un',2:'deux',3:'trois',4:'quatre',5:'cinq',6:'six',7:'sept',8:'huit',9:'neuf',10:'dix'}
DIZ = {20:'Vingt',30:'Trente',40:'Quarante',50:'Cinquante'}

def en_lettres(n):
    if n in LETTRES: return LETTRES[n].capitalize()
    d, u = (n//10)*10, n % 10
    if d in DIZ:
        return DIZ[d] if u == 0 else f"{DIZ[d]}-{LETTRES[u]}"
    return str(n)

def lire_sondages(chemin):
    S = []
    for ligne in io.open(chemin, encoding='utf-8').read().strip().split('\n')[1:]:
        c = ligne.split('|')
        if len(c) < 8: continue
        S.append(dict(d=c[0], f=c[1], n=int(c[2]), v=[float(x) for x in c[3:8]]))
    return sorted(S, key=lambda x: x['d'])

def lisser(S, ref, h=25.0):
    j = lambda s: (date(*map(int, s.split('-'))) - ref).days
    t = np.array([j(x['d']) for x in S], float)
    V = np.array([x['v'] for x in S]); n = np.array([x['n'] for x in S], float)
    grille = np.arange(t.min(), 1, 2.0)
    lisse = {p: [] for p in P}
    for g in grille:
        w = np.exp(-0.5*((t-g)/h)**2)*np.sqrt(n)
        X = np.column_stack([np.ones_like(t), t-g]); W = np.diag(w)
        b = np.linalg.solve(X.T@W@X, X.T@W@V)
        for k, p in enumerate(P): lisse[p].append(round(float(b[0][k]), 2))
    w = np.exp(-0.5*(t/h)**2)*np.sqrt(n)
    X = np.column_stack([np.ones_like(t), t]); W = np.diag(w)
    b = np.linalg.solve(X.T@W@X, X.T@W@V)
    pente = {p: round(float(b[1][k]*30), 1) for k, p in enumerate(P)}
    iso = lambda x: (ref + timedelta(days=int(x))).isoformat()
    return dict(grid=[iso(g) for g in grille], lisse=lisse,
                sondages=[{'d': x['d'], 'f': x['f'], 'n': x['n'], 'v': x['v']} for x in S],
                agregat={p: lisse[p][-1] for p in P}, pente=pente)

def recalculer(chemin):
    """openpyxl n'évalue pas les formules : on passe le classeur par LibreOffice."""
    import subprocess, os, tempfile, shutil
    if not shutil.which('soffice'):
        return chemin
    d = tempfile.mkdtemp()
    subprocess.run(['soffice','--headless','--convert-to','xlsx','--outdir',d,chemin],
                   capture_output=True, timeout=180)
    out = os.path.join(d, os.path.basename(chemin))
    return out if os.path.exists(out) else chemin

# Libellés courts affichés sur la page. Le classeur nomme les scénarios pour
# l'analyste (« +N2 », « Q2 ») ; la page s'adresse à un lecteur non spécialiste.
# Sans cette table, chaque passage du script écraserait les libellés rédigés.
LIBELLES = {
    'B – A + inertie locale':            'B – Avec inertie locale',
    'C – Fourchette haute (+N2)':        'C – Fourchette haute',
    'D – Fourchette basse (–N2)':        'D – Fourchette basse',
    'E – Amorti prop./uniforme (Q2)':    'E – Amorti prop./uniforme',
    'F – Régional + couche linguistique':'F – Régional + langue',
}

def lire_sieges(chemin):
    import openpyxl
    ws = openpyxl.load_workbook(chemin, data_only=True)['Sièges']
    out = {}
    for r in range(6, 20):
        nom = ws.cell(r, 1).value
        if not nom or not str(nom).startswith('Scén'): continue
        v = [ws.cell(r, c).value for c in range(2, 7)]
        if any(x is None for x in v): continue
        court = re.sub(r'^Scén\.\s*', '', str(nom)).replace('–', '–')
        court = LIBELLES.get(court, court)
        out[court] = dict(zip(['CAQ','PLQ','QS','PQ','PCQ'], [int(x) for x in v]))
    if not out: sys.exit("ERREUR : aucun scénario lu — le classeur n'a pas de valeurs calculées.\n         Relancer avec --recalculer, ou ouvrir puis enregistrer le fichier dans Excel.")
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sondages', default='data/sondages_national.csv')
    ap.add_argument('--classeur', required=True)
    ap.add_argument('--page', default='livrables/tendances.html')
    ap.add_argument('--autonome', default='tendances-quebec-2026.html',
                    help='copie autonome (document HTML complet) régénérée à chaque passage')
    ap.add_argument('--recalculer', action='store_true', help='recalculer le classeur via LibreOffice avant lecture')
    ap.add_argument('--date', help="date d'estimation AAAA-MM-JJ (défaut : aujourd'hui)")
    a = ap.parse_args()

    S = lire_sondages(a.sondages)
    ref = date(*map(int, a.date.split('-'))) if a.date else date.today()
    if (ref - date(*map(int, S[-1]['d'].split('-')))).days > 21:
        print(f"AVERTISSEMENT : le dernier sondage date du {S[-1]['d']}, soit plus de trois semaines "
              f"avant la date d'estimation. L'extrapolation devient hasardeuse.\n")
    D = lisser(S, ref)
    SE = lire_sieges(recalculer(a.classeur) if a.recalculer else a.classeur)
    page = io.open(a.page, encoding='utf-8').read()

    # état précédent, pour le rapport
    m = re.search(r'const D = (\{.*?\});\n', page, re.S)
    avant = json.loads(m.group(1)) if m else None

    page = re.sub(r'const D = \{.*?\};\n', 'const D = ' + json.dumps(D, ensure_ascii=False) + ';\n',
                  page, count=1, flags=re.S)
    ordre = ['PQ','PLQ','CAQ','PCQ','QS']
    corps = ',\n '.join("%r:{%s}" % (k, ','.join(f"{p}:{v[p]}" for p in ordre)) for k, v in SE.items())
    page = re.sub(r'const SEATS=\{.*?\};\n', 'const SEATS={\n ' + corps + '};\n',
                  page, count=1, flags=re.S).replace("'", "'")
    jour = f"{ref.day}{'er' if ref.day==1 else ''} {MOIS[ref.month-1]} {ref.year}"
    page = re.sub(r'(lissage local · )[^<]+(</p>)', r'\g<1>' + jour + r'\2', page, count=1)
    page = re.sub(r'^[A-ZÉÈ][a-zé-]+(-[a-z]+)? sondages depuis', en_lettres(len(S)) + ' sondages depuis',
                  page, count=1, flags=re.M)

    io.open(a.page, 'w', encoding='utf-8').write(page)
    tete, corps_html = page.split('<div class="wrap">', 1)
    io.open(a.autonome, 'w', encoding='utf-8').write(
        '<!doctype html><html lang="fr"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        '<style>:root{color-scheme:light dark}body{margin:0}img{max-width:100%}</style>'
        + tete + '</head><body><div class="wrap">' + corps_html + '</body></html>')

    print(f"Page mise à jour : {len(S)} sondages, estimation au {jour}\n")
    print(f"{'':6s} {'niveau':>8s} {'pente/30j':>10s}" + ("   (précédent)" if avant else ""))
    for p in P:
        prec = f"   {avant['agregat'][p]:.1f} / {avant['pente'][p]:+.1f}" if avant else ""
        print(f"{p:6s} {D['agregat'][p]:8.1f} {D['pente'][p]:+10.1f}{prec}")
    print("\nSièges :")
    for k, v in SE.items(): print(f"  {k:42s} " + "  ".join(f"{p} {v[p]:3d}" for p in ordre))

    print("\n--- PHRASES À VÉRIFIER À LA MAIN ---")
    alertes = []
    if avant:
        for p in P:
            if abs(D['pente'][p] - avant['pente'][p]) >= 1.0:
                alertes.append(f"la pente du {p} passe de {avant['pente'][p]:+.1f} à {D['pente'][p]:+.1f} — "
                               f"le texte de « Ce que la courbe dit » cite peut-être l'ancienne valeur")
            if np.sign(D['pente'][p]) != np.sign(avant['pente'][p]):
                alertes.append(f"le {p} CHANGE DE DIRECTION ({avant['pente'][p]:+.1f} → {D['pente'][p]:+.1f}) — "
                               f"le récit de la page est probablement faux")
    maxi = max(max(v.values()) for v in SE.values())
    if maxi >= 64:
        alertes.append(f"UN PARTI ATTEINT LA MAJORITÉ ({maxi} sièges) — la page affirme partout qu'aucun scénario n'y arrive")
    for m2 in re.finditer(r'\b(\d{1,2},\d)\s*(?:points?|%)', page):
        pass
    print("\n".join("  ⚠ " + x for x in alertes) if alertes else "  aucune alerte automatique")
    print("""
  À relire dans tous les cas, ces passages contiennent des chiffres écrits en toutes lettres :
    · le chapeau (« La CAQ a regagné une quinzaine de points… »)
    · les trois paragraphes de « Ce que la courbe dit »
    · le paragraphe PCQ de « D'ici le 5 octobre » et les circonscriptions citées
    · l'encadré « quatorze circonscriptions sur soixante-huit »
    · les tableaux des deux cartes « fractures »
    · la ligne de mise à jour du pied de page""")

if __name__ == '__main__':
    main()
